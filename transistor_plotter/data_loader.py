from __future__ import annotations

import re
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from .models import Curve, CurveSet, DeviceCurves, SensorFiles

FILE_RE = re.compile(r"^(DIODE|IV|TRANS)_(.+?)_4F_?50_T_296_K\.xlsx$", re.IGNORECASE)


def discover_sensors(data_dir: Path) -> list[SensorFiles]:
    grouped: dict[tuple[str, str], dict[str, Path]] = {}
    for path in data_dir.rglob("*.xlsx"):
        match = FILE_RE.match(path.name)
        if not match:
            continue
        kind, device_id = match.group(1).upper(), match.group(2)
        key = (path.parent.name, device_id)
        grouped.setdefault(key, {})[kind] = path

    sensors: list[SensorFiles] = []
    for (group, device_id), files in grouped.items():
        if {"DIODE", "IV", "TRANS"}.issubset(files):
            sensors.append(
                SensorFiles(
                    group=group,
                    device_id=device_id,
                    diode=files["DIODE"],
                    iv=files["IV"],
                    trans=files["TRANS"],
                )
            )

    return sorted(sensors, key=lambda item: (item.group, item.device_id))


def load_device_curves(sensor: SensorFiles) -> DeviceCurves:
    return DeviceCurves(
        sensor=sensor,
        diode_ig_vgs=_load_curve_set(
            sensor.diode,
            sheet_index=0,
            title="Gate leakage",
            xlabel="Vgs (V)",
            ylabel="Ig (uA/mm)",
        ),
        iv_id_vds=_load_curve_set(
            sensor.iv,
            sheet_index=0,
            title="Output characteristic",
            xlabel="Vds (V)",
            ylabel="Id (mA/mm)",
        ),
        trans_id_vgs=_load_curve_set(
            sensor.trans,
            sheet_index=0,
            title="Transfer characteristic",
            xlabel="Vgs (V)",
            ylabel="Id (mA/mm)",
        ),
        trans_gm_vgs=_load_curve_set(
            sensor.trans,
            sheet_index=1,
            title="Transconductance",
            xlabel="Vgs (V)",
            ylabel="gm (mS/mm)",
        ),
    )


def _load_curve_set(path: Path, sheet_index: int, title: str, xlabel: str, ylabel: str) -> CurveSet:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[sheet_index]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return CurveSet(title=title, xlabel=xlabel, ylabel=ylabel, curves=())

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    x_values = [_to_float(row[0]) for row in rows[1:] if row and _to_float(row[0]) is not None]
    x = np.asarray(x_values, dtype=float)

    curves: list[Curve] = []
    for column_index, header in enumerate(headers[1:], start=1):
        if not header:
            continue

        y_values: list[float] = []
        paired_x: list[float] = []
        for row in rows[1:]:
            if column_index >= len(row):
                continue
            x_value = _to_float(row[0])
            y_value = _to_float(row[column_index])
            if x_value is None or y_value is None:
                continue
            paired_x.append(x_value)
            y_values.append(y_value)

        if y_values:
            curves.append(
                Curve(
                    label=header,
                    x=np.asarray(paired_x, dtype=float) if len(paired_x) != len(x) else x,
                    y=np.asarray(y_values, dtype=float),
                )
            )

    return CurveSet(title=title, xlabel=xlabel, ylabel=ylabel, curves=tuple(curves))


def _to_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
